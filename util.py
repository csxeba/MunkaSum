import datetime

from MunkaSum.time import Time

exceptions = {"Bogdánné Major Andrea": [-30] * 5,
              "Simon Ildikó": [-30] * 5,
              "Gherghely Ildikó": [0, 0, 0, 90, 0]}

INFLTYPE = "xlsx"
OUTFLTYPE = "xlsx"
GUI = True


def parse_csv(path: str):
    with open(path) as f:
        text = f.read()
        sep = "\t" if "\t" in text else ";"
        f.close()

    lines = []
    for line in text.split("\n"):
        line = line.split(sep)
        if any(line):
            lines.append(line)

    if lines[-1][0][:4] == "Össz":
        lines = lines[:-1]

    header = fix_header(lines[0])
    lines = lines[1:]
    matrix = extract_matrix(lines)
    return matrix, header


def parse_xl(path: str):
    import openpyxl as xl
    wb = xl.load_workbook(path, data_only=True)
    sheets = [sheet for sheet in wb if sheet.dimensions != "A1:A1"]
    if not sheets:
        raise RuntimeError("No valid sheets in file!")

    rows = []
    for sheet in sheets:
        rows += list(sheet.rows)
    lines = [[cell.value for cell in row] for row in rows]
    lines = [row for row in lines if any(row)]
    header = fix_header(lines[0])
    matrix = extract_matrix(lines)

    return matrix, header


def fix_header(raw_header):
    assert len(raw_header) == 8 and raw_header[0] == "Név"
    header = raw_header[:6]
    header[3] = "érk_hiba"
    header[5] = "táv_hiba"
    return header


def extract_matrix(lines):
    matrix = []
    for line in lines:
        if (line[0] == "Név") or \
                ("összesen" in line[0].lower()) or \
                not line[0]:
            print("Skipping invalid line")
            continue
        name, date, arrive, depart = line[0], todate(line[1]), totime(line[3]), totime(line[5])
        if arrive == "n.a.":
            arrive = totime(line[2])
            if arrive == "n.a.":
                arrive = "n.a."
        if depart == "n.a.":
            depart = totime(line[4])
            if depart == "n.a.":
                depart = "n.a."
        arr_err, dep_err = assert_line([name, date, arrive, depart])
        matrix.append([name, date, arrive, arr_err, depart, dep_err])

    return matrix


def summarize(matrix):
    names = sorted(list(set([line[0] for line in matrix])))
    dictionary = {name: [0, 0, 0, 0] for name in names}
    for line in matrix:
        if line[3].epochs > 0 and line[3] != "n.a.":
            dictionary[line[0]][0] += line[3].epochs
            dictionary[line[0]][1] += 1
        if line[5].epochs > 0 and line[5] != "n.a.":
            dictionary[line[0]][2] += line[5].epochs
            dictionary[line[0]][3] += 1
    return dictionary, names


def todate(x):
    if isinstance(x, str):
        nums = [int(d) for d in x.split(".")]
    else:
        nums = [x.year, x.month, x.day]
    return datetime.date(year=nums[0], month=nums[1], day=nums[2])


def totime(x):
    if isinstance(x, str):
        if x == "n.a.":
            return "n.a."
        nums = [int(d) for d in x.split(":")]
    else:
        nums = [x.hour, x.minute]
    return Time(hours=nums[0], minutes=nums[1])


def assert_line(line):

    def calc_patience(nm):
        pt = exceptions[nm][date.weekday()] if nm in exceptions else 0
        if pt < 0:
            return Time.from_epochs(abs(pt), sign=-1)
        else:
            return Time.from_epochs(pt)

    name, date, arrive, depart = line
    if arrive == "n.a." or depart == "n.a.":
        return ["n.a.", "n.a."]
    patience = calc_patience(name)
    result = [arrive - (Time(hours=7, minutes=30) + patience)]

    if date.weekday() == 4:
        referece = Time(hours=13, minutes=30) + patience
    else:
        referece = Time(hours=16, minutes=0) + patience

    result.append(referece - depart)
    return result


def tk_get_path():
    from tkinter import Tk
    from tkinter.filedialog import askopenfilename

    tk = Tk()
    tk.withdraw()
    inpath = askopenfilename()
    tk.destroy()

    return inpath


def dump_to_csv(matrix, headers, outroot):
    outchain = "\t".join(headers) + "\n"
    for line in matrix:
        outchain += "\t".join([str(e) for e in line[:3]]) + "\t"
        outchain += str(line[3].epochs) + "\t"
        outchain += str(line[4]) + "\t" + str(line[5].epochs) + "\n"
    with open(outroot + "jelolt.csv", "w") as outfl:
        outfl.write(outchain)
        outfl.close()
    # print("-" * 70)
    # print(outchain)

    dictionary, names = summarize(matrix)

    outchain = "Név\tKésés\tDarab\tKorai_indulás\tDarab\n"
    for name in names:
        outchain += name + "\t" + "\t".join([str(element) for element in dictionary[name]]) + "\n"
    with open(outroot + "osszesites.csv", "w") as outfl:
        outfl.write(outchain)
        outfl.close()
    # print("-" * 70)
    # print(outchain)


def dump_to_xl(matrix, headers, outpath):
    from openpyxl import Workbook

    wb = Workbook()
    wb.create_sheet("Jelzett")
    wb.create_sheet("Összesített")
    wb.remove("Sheet")

    ws = wb.get_sheet_by_name("Jelzett")
    ws.append(headers)

    for i, row in enumerate(range(2, len(matrix))):
        arep = matrix[i][3].epochs
        drep = matrix[i][5].epochs
        ws["A" + str(row)].value = matrix[i][0]  # Name
        ws["B" + str(row)].value = str(matrix[i][1])  # Date
        ws["C" + str(row)].value = str(matrix[i][2])  # Arrive
        ws["D" + str(row)].value = arep if arep > 0 else 0
        ws["E" + str(row)].value = str(matrix[i][4])  # Depart
        ws["F" + str(row)].value = drep if drep > 0 else 0

    dictionary, names = summarize(matrix)

    summary = []
    for name, (lates, nlates, earlies, nealries) in dictionary.items():
        summary.append((name, lates, nlates, earlies, nealries))

    ws = wb.get_sheet_by_name("Összesített")
    ws.append("Név,Késés,Darab,Korai_indulás,Darab".split(","))
    for row in summary:
        ws.append(row)
    wb.save(outpath)


def main():
    if GUI:
        path = tk_get_path()
    else:
        path = "input" + "." + INFLTYPE

    parser = parse_csv if INFLTYPE == "csv" else parse_xl
    matrix, header = parser(path)
    if OUTFLTYPE == "csv":
        dump_to_csv(matrix, header, outroot="E:/tmp/")
    else:
        dump_to_xl(matrix, header, outpath="E:/tmp/output.xlsx")


if __name__ == '__main__':
    main()
