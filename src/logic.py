import datetime
import warnings

import openpyxl as xl
from .time import Time, WorkingSaturday

exceptions = {
    "Bogdánné Major Andrea": (
        [-30, -30, -30, -30, -30],
        [-30, -30, -30, -30, -30]),
    "Simon Ildikó": (
        [-30, -30, -30, -30, -30],
        [-30, -30, -30, -30, -30]),
    "Gherghely Ildikó": (
        [45, 45, 45, 45, 45],
        [-15, 60, -15, 60, 135])
}

ideal_header = ["Név", "Dátum", "Érkezés", "", "Távozás", "",
                "Napi rögzítési eltérés", "Rögzítette (TASZ, dátum)"]


working_saturdays = {
    "2016.10.25": 4,
}


def parse_csv(path: str):
    with open(path) as f:
        text = f.read()
        sep = "\t" if "\t" in text else ";"
        f.close()

    lines = []
    for line in text.split("\n"):
        line = line.split(sep)
        if any(line):
            lines.append(line)

    if lines[-1][0][:4] == "Össz":
        lines = lines[:-1]

    header = fix_header(lines[0])
    lines = lines[1:]
    matrix = extract_matrix(lines)
    return matrix, header


def parse_xl(path: str):
    wb = xl.load_workbook(path, data_only=True)
    sheets = [sheet for sheet in wb if validate(sheet)]
    if not sheets:
        raise RuntimeError("No valid sheets in file!")

    rows = []
    for sheet in sheets:
        rows += list(sheet.rows)
    lines = [[cell.value for cell in row] for row in rows]
    lines = [row for row in lines if any(row)]
    header = fix_header(lines[0])
    matrix = extract_matrix(lines)

    return matrix, header


def validate(sheet):
    if sheet.dimensions == "A1:A1":
        warnings.warn("Empty sheet found!", RuntimeWarning)
        return False
    if sheet.dimensions[:4] != "A1:H":
        warnings.warn("Invalid data!", RuntimeWarning)
        return False
    header = [cells[0].value for cells in sheet["A1:H1"]]
    if not all([left == right for left, right in zip(header, ideal_header)]):
        warnings.warn("Possibly missing header!", RuntimeWarning)
    return True


def fix_header(raw_header):
    assert len(raw_header) == 8 and raw_header[0] == "Név"
    if not all([left == right for left, right in zip(raw_header, ideal_header)]):
        raw_header = ideal_header

    header = raw_header[:6]
    header[3] = "érk_hiba"
    header[5] = "táv_hiba"
    return header


def extract_matrix(lines):
    matrix = []
    for line in lines:
        if (line[0] == "Név") or \
                ("összesen" in line[0].lower()) or \
                not line[0]:
            print("Skipping invalid line")
            continue
        name, date, arrive, depart = line[0], todate(line[1]), totime(line[3]), totime(line[5])
        if arrive == "n.a.":
            arrive = totime(line[2])
            if arrive == "n.a.":
                arrive = "n.a."
        if depart == "n.a.":
            depart = totime(line[4])
            if depart == "n.a.":
                depart = "n.a."
        arr_err, dep_err = assert_line([name, date, arrive, depart])
        matrix.append([name, date, arrive, arr_err, depart, dep_err])

    return matrix


def summarize(matrix):
    names = sorted(list(set([line[0] for line in matrix])))
    # m_over, m_late, m_nover, m_nlate, e_over, e_late, e_nover, e_nlate
    dictionary = {name: [0, 0, 0, 0, 0, 0, 0, 0] for name in names}
    for line in matrix:
        if line[3] != "n.a.":
            if line[3].epochs < 0:  # MORNING OVERWORK
                dictionary[line[0]][0] += line[3].epochs
                dictionary[line[0]][2] += 1
            else:  # MORNING UNDERWORK
                dictionary[line[0]][1] += line[3].epochs
                dictionary[line[0]][3] += 1
        if line[5] != "n.a.":  # EVENING OVERWORK
            if line[5].epochs < 0:
                dictionary[line[0]][4] += line[5].epochs
                dictionary[line[0]][6] += 1
            else:  # EVENING UNDERWORK
                dictionary[line[0]][5] += line[5].epochs
                dictionary[line[0]][7] += 1
    return dictionary, names


def todate(x):
    if isinstance(x, str):
        nums = [int(d) for d in x.split(".")]
    else:
        nums = [x.year, x.month, x.day]
    if x in working_saturdays:
        return WorkingSaturday(year=nums[0], month=nums[1], day=nums[2], schedule=working_saturdays[x])
    else:
        return datetime.date(year=nums[0], month=nums[1], day=nums[2])


def totime(x):
    if isinstance(x, str):
        if x == "n.a.":
            return "n.a."
        nums = [int(d) for d in x.split(":")]
    else:
        nums = [x.hour, x.minute]
    return Time(hours=nums[0], minutes=nums[1])


def assert_line(line):

    def calc_patience(nm, evening):
        pt = exceptions[nm][int(evening)][date.weekday()] if nm in exceptions else 0
        if pt < 0:
            return Time.from_epochs(abs(pt), sign=-1)
        else:
            return Time.from_epochs(pt)

    name, date, arrive, depart = line
    if arrive == "n.a." or depart == "n.a.":
        return ["n.a.", "n.a."]
    morning_patience = calc_patience(name, evening=False)
    evening_patience = calc_patience(name, evening=True)
    result = [arrive - (Time(hours=7, minutes=30) + morning_patience)]

    if date.weekday() == 4:
        reference = Time(hours=13, minutes=30) - evening_patience
    else:
        reference = Time(hours=16, minutes=0) - evening_patience

    result.append(reference - depart)
    return result


def dump_to_csv(matrix, headers, outroot):

    def build_big_data_string():
        outchain = "\t".join(headers) + "\n"
        for line in matrix:
            arr_err = line[3].epochs
            dep_err = line[5].epochs
            outchain += "\t".join([str(e) for e in line[:3]]) + "\t"
            outchain += str(arr_err if arr_err > 0 else "") + "\t"
            outchain += str(line[4]) + "\t" + str(dep_err if dep_err > 0 else "") + "\n"

        return outchain, "jelolt.csv"

    def build_summary_data_string():
        dictionary, names = summarize(matrix)

        outchain = "Név\tReggel_túlóra\tKésés\ttúlóra_darab\tkésés_darab\t"
        outchain += "Esti_túlóra\tKorai_indulás\ttúlóra_darab\tkésés_darab\n"
        for name in names:
            outchain += name + "\t" + "\t".join([str(element) for element in dictionary[name]]) + "\n"
        return outchain, "osszesitett.csv"

    def dump_chain(chain, flname):
        with open(outroot + flname, "w") as outfl:
            outfl.write(chain)
            outfl.close()

    dump_chain(*build_big_data_string())
    dump_chain(*build_summary_data_string())


def dump_to_xl(matrix, headers, outpath):
    from openpyxl import Workbook

    def build_big_worksheet():
        ws = wb.get_sheet_by_name("Jelzett")
        ws.append(headers)

        for i, row in enumerate(range(2, len(matrix))):
            arep = matrix[i][3].epochs
            drep = matrix[i][5].epochs
            ws["A" + str(row)].value = matrix[i][0]  # Name
            ws["B" + str(row)].value = str(matrix[i][1])  # Date
            ws["C" + str(row)].value = str(matrix[i][2])  # Arrive
            ws["D" + str(row)].value = arep if arep > 0 else ""
            ws["E" + str(row)].value = str(matrix[i][4])  # Depart
            ws["F" + str(row)].value = drep if drep > 0 else ""

    def build_summary_worksheet():
        # dict: m_over, m_late, m_nover, m_nlate, e_over, e_late, e_nover, e_nlate
        dictionary, names = summarize(matrix)

        summary = []
        names = sorted(list(dictionary.keys()))
        for name in names:
            data = dictionary[name]
            summary.append([name] + data + [data[0] + data[1] + data[4] + data[5]])

        ws = wb.get_sheet_by_name("Összesített")
        header = "név, reggeli_túlóra, késés, túlóra_darab, késés_darab, "
        header += "esti_túlóra, korai_indulás, túlóra_darab, ki_darab, perc_összes"
        ws.append(header.split(", "))
        for row in summary:
            ws.append(row)

    def remove_orphan_sheet():
        sheets = list(wb.get_sheet_names())
        if "Sheet" in sheets:
            wb.remove(wb.get_sheet_by_name("Sheet"))
            sheets.remove("Sheet")

    wb = Workbook()
    wb.create_sheet("Jelzett")
    wb.create_sheet("Összesített")

    build_big_worksheet()
    build_summary_worksheet()
    remove_orphan_sheet()

    wb.save(outpath)
